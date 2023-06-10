import openpyxl
import pandas
from openpyxl.utils.dataframe import dataframe_to_rows

#从更新的大表姐中提取dataframe
file_path_newData = r'C:\Users\he kelly\Downloads\GL Dump query Jun.16 _ with RU and Currency Conversion TW 0503.xlsx'
wbNewData = openpyxl.load_workbook(file_path_newData)
wsNewData = wbNewData.worksheets[0]
newData = wsNewData.values
newDataColumns = wsNewData.max_column
newDataCols = []
for i in range(1, newDataColumns+1):
    cellvalue = wsNewData.cell(row=2, column=i).value
    newDataCols.append(cellvalue)
newData = list(newData)
newData = newData[2:]
dfNewData = pandas.DataFrame(newData, columns=newDataCols)
dfNewData.dropna(how='all', axis=1, inplace=True)
print(dfNewData)


#提取原文件中的dataframe
file_path_Data = r'C:\Users\he kelly\Desktop\OC\2023\2023.3\SH.xlsx'
wbData = openpyxl.load_workbook(file_path_Data)
wsData = wbData["Data"]
data = wsData.values
dataColumns = wsData.max_column
dataCols = []
for i in range(1, dataColumns+1):
    cellvalue = wsData.cell(row=2, column=i).value
    dataCols.append(cellvalue)
data = list(data)
data = data[2:]
dfData = pandas.DataFrame(data, columns=dataCols)
dfData.dropna(how='all', axis=1, inplace=True)

#对齐两个dataframe的列数

for i in dfData.columns:
    print(i)
    if i in dfNewData.columns:
        print("Yes")
    else:
        print("NO")
        dfData.drop(f'{i}', axis=1, inplace=True)

#截取需要更新的数据部分
# max = dfData['JH Created Date'].max()
# dfNewData_filtered = dfNewData.loc[dfNewData['JH Created Date'] > max, :]


intersected_df = pandas.merge(dfNewData, dfData, how='inner')
dfNewData_filtered = pandas.concat([dfNewData, intersected_df, intersected_df]).drop_duplicates(keep=False)


#将需更新部分贴进excel
for r in dataframe_to_rows(dfNewData_filtered,index=False, header=False):
    wsData.append(r)
wbData.save(r'C:\Users\he kelly\Desktop\OC\2023\2023.3\SH1.xlsx')

#Check
sumNewData = dfNewData.loc[dfNewData['Period Name']=='MAR-23','Amount Func Cur'].sum()+dfNewData.loc[dfNewData['Period Name']=='APR-23','Amount Func Cur'].sum()
sumData = dfData.loc[dfData['Period Name']=='MAR-23','Amount Func Cur'].sum()+dfData.loc[dfData['Period Name']=='APR-23','Amount Func Cur'].sum()+dfNewData_filtered['Amount Func Cur'].sum()
if sumNewData == sumData:
    print("Checked")
else:
    print("Error")


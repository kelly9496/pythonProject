
import openpyxl
file_path = r"C:\Users\he kelly\Desktop\OC\printing fees.xlsx"
workbook = openpyxl.load_workbook(file_path)
print(workbook.sheetnames)
sheet = workbook["SH"]
list = sheet["AB"]
print(list[1].value)

record = []

for aIndex in range(1,len(list)):
    if aIndex in record:
        continue
    #print(list[aIndex])
    for bIndex in range(1,len(list)):
        #print(list[bIndex])
        if bIndex in record:
            continue
        if (float(list[aIndex].value)+float(list[bIndex].value))==0:
            record.append(aIndex)
            record.append(bIndex)
            break

print(record)


for i in record:
        cell = list[i]
        cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='FFFF50')
workbook.save(file_path)